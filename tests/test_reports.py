"""Tests: reports with filters, date validation, PDF/Excel export (task 13.9)"""

from io import BytesIO
from datetime import datetime, timezone, timedelta
from decimal import Decimal
from zoneinfo import ZoneInfo

from app.core.config import settings

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession


def _now():
    return datetime.now(timezone.utc)


def _range(days_back=1):
    to = _now()
    frm = to - timedelta(days=days_back)
    return frm.isoformat(), to.isoformat()


@pytest_asyncio.fixture(autouse=True)
async def _company(company_config):
    """All report tests require a configured company."""
    return company_config


async def _seed_ingreso(client, admin_token, operator_token):
    cat = await client.post(
        "/api/v1/categories",
        json={"name": "ReportCat"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    cat_id = cat.json()["id"]
    prod = await client.post(
        "/api/v1/products",
        json={"name": "ReportProd", "category_id": cat_id, "pvp": "1.00"},
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    prod_id = prod.json()["id"]
    await client.post(
        "/api/v1/inventory/ingresos",
        json={"lines": [{"product_id": prod_id, "quantity": "5", "unit_cost": "2.00"}]},
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    return prod_id


async def _seed_egreso(client, admin_token, operator_token, egreso_type="sale"):
    prod_id = await _seed_ingreso(client, admin_token, operator_token)
    payload = {
        "egreso_type": egreso_type,
        "lines": [{"product_id": prod_id, "quantity": "1", "unit_price": "3.00"}],
    }
    if egreso_type == "sale":
        payload["seller_name"] = "VENDEDOR TEST"
    await client.post(
        "/api/v1/inventory/egresos",
        json=payload,
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    return prod_id


@pytest.mark.asyncio
async def test_report_ingresos_json(
    client: AsyncClient, admin_token: str, operator_token: str
):
    await _seed_ingreso(client, admin_token, operator_token)
    frm, to = _range(days_back=1)
    resp = await client.get(
        "/api/v1/reports/ingresos",
        params={"date_from": frm, "date_to": to, "format": "json"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert isinstance(data, list)
    assert len(data) >= 1
    assert "number" in data[0]


@pytest.mark.asyncio
async def test_report_ingresos_json_date_only_includes_same_day(
    client: AsyncClient, admin_token: str, operator_token: str
):
    await _seed_ingreso(client, admin_token, operator_token)
    # Filter by the APP_TIMEZONE local day (matches how the API resolves
    # date-only bounds), so the test is robust to the UTC/local-day boundary.
    today = _now().astimezone(ZoneInfo(settings.APP_TIMEZONE)).date().isoformat()
    resp = await client.get(
        "/api/v1/reports/ingresos",
        params={"date_from": today, "date_to": today, "format": "json"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert isinstance(data, list)
    assert len(data) >= 1


@pytest.mark.asyncio
async def test_report_ingresos_filters_by_type(
    client: AsyncClient, admin_token: str, operator_token: str
):
    prod_id = await _seed_ingreso(client, admin_token, operator_token)
    await client.post(
        "/api/v1/inventory/ingresos",
        json={
            "ingreso_type": "production",
            "lines": [{"product_id": prod_id, "quantity": "2", "unit_cost": "2.00"}],
        },
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    frm, to = _range(days_back=1)
    resp = await client.get(
        "/api/v1/reports/ingresos",
        params={"date_from": frm, "date_to": to, "format": "json", "type": "production"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 1


@pytest.mark.asyncio
async def test_report_egresos_filters_by_type(
    client: AsyncClient, admin_token: str, operator_token: str
):
    await _seed_egreso(client, admin_token, operator_token, egreso_type="sale")
    await _seed_egreso(
        client,
        admin_token,
        operator_token,
        egreso_type="internal_consumption",
    )
    frm, to = _range(days_back=1)
    resp = await client.get(
        "/api/v1/reports/egresos",
        params={
            "date_from": frm,
            "date_to": to,
            "format": "json",
            "type": "internal_consumption",
        },
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 1


@pytest.mark.asyncio
async def test_report_ingresos_excel(
    client: AsyncClient, admin_token: str, operator_token: str
):
    await _seed_ingreso(client, admin_token, operator_token)
    frm, to = _range(days_back=1)
    resp = await client.get(
        "/api/v1/reports/ingresos",
        params={"date_from": frm, "date_to": to, "format": "excel"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]


@pytest.mark.asyncio
async def test_report_ingresos_pdf(
    client: AsyncClient, admin_token: str, operator_token: str
):
    await _seed_ingreso(client, admin_token, operator_token)
    frm, to = _range(days_back=1)
    resp = await client.get(
        "/api/v1/reports/ingresos",
        params={"date_from": frm, "date_to": to, "format": "pdf"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"


@pytest.mark.asyncio
async def test_report_ingresos_inverted_dates(client: AsyncClient, admin_token: str):
    to, frm = _range(days_back=1)  # intentionally swapped
    resp = await client.get(
        "/api/v1/reports/ingresos",
        params={"date_from": frm, "date_to": to},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 422
    assert resp.json()["code"] == "INVALID_DATE_RANGE"


@pytest.mark.asyncio
async def test_report_stock_json(
    client: AsyncClient, admin_token: str, operator_token: str
):
    await _seed_ingreso(client, admin_token, operator_token)
    resp = await client.get(
        "/api/v1/reports/stock?format=json",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert isinstance(data, list)
    assert len(data) >= 1
    assert "stock_actual" in data[0]
    assert "bajo_stock" in data[0]


@pytest.mark.asyncio
async def test_report_stock_excel(client: AsyncClient, admin_token: str):
    resp = await client.get(
        "/api/v1/reports/stock?format=excel",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]


@pytest.mark.asyncio
async def test_report_stock_excel_hides_company_header_when_logo_disabled(
    client: AsyncClient,
    admin_token: str,
    db_session: AsyncSession,
):
    from app.models.system_param import SystemParam

    db_session.add(
        SystemParam(key="report_include_logo", value="false", description="test")
    )
    await db_session.commit()

    resp = await client.get(
        "/api/v1/reports/stock?format=excel",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws["A1"].value == "Reporte de Stock Actual"


@pytest.mark.asyncio
async def test_report_consolidado(
    client: AsyncClient, admin_token: str, operator_token: str
):
    await _seed_ingreso(client, admin_token, operator_token)
    frm, to = _range(days_back=1)
    resp = await client.get(
        "/api/v1/reports/consolidado",
        params={"date_from": frm, "date_to": to},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert "movements" in data
    assert "active_products" in data
    assert "products_below_minimum" in data
    assert data["movements"]["IN"] >= 1


@pytest.mark.asyncio
async def test_report_stock_valorizado(
    client: AsyncClient, admin_token: str, operator_token: str
):
    await _seed_ingreso(client, admin_token, operator_token)
    resp = await client.get(
        "/api/v1/reports/stock-valorizado",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert "items" in data
    assert "total_value" in data
    assert "method" in data


@pytest.mark.asyncio
async def test_report_stock_valorizado_matches_kardex_balance_value_weighted_average(
    client: AsyncClient,
    admin_token: str,
    operator_token: str,
    db_session: AsyncSession,
):
    from app.models.system_param import SystemParam

    db_session.add(
        SystemParam(key="kardex_method", value="WEIGHTED_AVERAGE", description="test")
    )
    await db_session.commit()

    cat = await client.post(
        "/api/v1/categories",
        json={"name": "ReportWeightedCat"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    cat_id = cat.json()["id"]
    prod = await client.post(
        "/api/v1/products",
        json={"name": "ReportWeightedProd", "category_id": cat_id, "pvp": "1.00"},
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    prod_id = prod.json()["id"]

    await client.post(
        "/api/v1/inventory/ingresos",
        json={
            "lines": [{"product_id": prod_id, "quantity": "10", "unit_cost": "10.00"}]
        },
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    await client.post(
        "/api/v1/inventory/ingresos",
        json={
            "lines": [{"product_id": prod_id, "quantity": "10", "unit_cost": "20.00"}]
        },
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    await client.post(
        "/api/v1/inventory/egresos",
        json={
            "seller_name": "VENDEDOR TEST",
            "lines": [{"product_id": prod_id, "quantity": "5", "unit_price": "30.00"}]
        },
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    kardex_resp = await client.get(
        f"/api/v1/kardex/{prod_id}",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert kardex_resp.status_code == 200
    closing_value = Decimal(str(kardex_resp.json()["closing_balance_value"]))

    report_resp = await client.get(
        "/api/v1/reports/stock-valorizado",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert report_resp.status_code == 200
    items = report_resp.json()["items"]
    item = next(i for i in items if i["id"] == prod_id)
    report_value = Decimal(str(item["value"]))

    assert report_value == closing_value


@pytest.mark.asyncio
async def test_report_requires_admin_or_supervisor(
    client: AsyncClient, operator_token: str
):
    frm, to = _range(days_back=1)
    resp = await client.get(
        "/api/v1/reports/ingresos",
        params={"date_from": frm, "date_to": to},
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    assert resp.status_code == 403
