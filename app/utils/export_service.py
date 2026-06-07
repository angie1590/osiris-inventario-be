"""Export utilities for PDF and Excel report generation."""

import io
from datetime import datetime
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExportService:
    @staticmethod
    def to_pdf(
        headers: list[str],
        rows: list[list[Any]],
        title: str,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        generated_by: str = "",
        company_header: dict[str, Any] | None = None,
    ) -> bytes:
        section_rows: list[int] = []
        subtotal_rows: list[int] = []

        numeric_header_hints = {
            "monto",
            "saldo",
            "total",
            "costo",
            "precio",
            "valor",
            "stock",
            "cantidad",
            "entrada",
            "salida",
            "pvp",
        }

        def _is_numeric_value(value: str) -> bool:
            cleaned = value.strip()
            if not cleaned:
                return False
            cleaned = cleaned.replace("$", "").replace(" ", "")
            cleaned = cleaned.replace(".", "").replace(",", ".")
            try:
                float(cleaned)
                return True
            except ValueError:
                return False

        def _normalize_row(row: list[Any], columns: int, current_row: int) -> list[str]:
            normalized = [str(cell) for cell in row]
            if not normalized:
                normalized = [""]

            first = normalized[0]
            if first.startswith("__SECTION__:"):
                normalized[0] = first.replace("__SECTION__:", "", 1).strip()
                section_rows.append(current_row)
            elif first.startswith("__SUBTOTAL__:"):
                normalized[0] = first.replace("__SUBTOTAL__:", "", 1).strip()
                subtotal_rows.append(current_row)

            if len(normalized) < columns:
                normalized.extend([""] * (columns - len(normalized)))
            elif len(normalized) > columns:
                normalized = normalized[:columns]
            return normalized

        def _draw_page_footer(canvas, doc):
            canvas.saveState()
            canvas.setFont("Helvetica", 8)
            canvas.setFillColor(colors.HexColor("#4B5563"))
            canvas.drawRightString(
                doc.pagesize[0] - doc.rightMargin,
                8 * mm,
                f"Página {canvas.getPageNumber()}",
            )
            canvas.restoreState()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            topMargin=18 * mm,
            bottomMargin=14 * mm,
            leftMargin=10 * mm,
            rightMargin=10 * mm,
        )
        styles = getSampleStyleSheet()
        styles.add(
            ParagraphStyle(
                name="MetaInfo",
                parent=styles["Normal"],
                fontName="Helvetica",
                fontSize=8,
                leading=10,
                textColor=colors.HexColor("#374151"),
            )
        )
        styles.add(
            ParagraphStyle(
                name="MetaInfoRight",
                parent=styles["MetaInfo"],
                alignment=TA_RIGHT,
            )
        )
        story = []

        generated_label = datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")

        # Corporate header
        if company_header:
            razon = company_header.get("razon_social", "")
            nombre = company_header.get("nombre_comercial")
            ruc = company_header.get("ruc", "")
            story.append(Paragraph(f"<b>{razon}</b>", styles["Heading2"]))
            if nombre:
                story.append(Paragraph(nombre, styles["MetaInfo"]))
            story.append(Paragraph(f"RUC: {ruc}", styles["MetaInfo"]))
            story.append(Spacer(1, 2 * mm))

        # Header section
        story.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
        story.append(Spacer(1, 1.5 * mm))

        meta_left = []
        if date_from and date_to:
            meta_left.append(
                f"Período: {date_from.strftime('%d/%m/%Y')} - {date_to.strftime('%d/%m/%Y')}"
            )
        meta_left.append(f"Usuario: {generated_by}")
        meta_table = Table(
            [
                [
                    Paragraph("<br/>".join(meta_left), styles["MetaInfo"]),
                    Paragraph(f"Emitido: {generated_label}", styles["MetaInfoRight"]),
                ]
            ],
            colWidths=[doc.width * 0.7, doc.width * 0.3],
        )
        meta_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        story.append(meta_table)
        story.append(Spacer(1, 3 * mm))

        # Data table
        table_data = [headers]
        for idx, row in enumerate(rows, start=1):
            table_data.append(_normalize_row(row, len(headers), idx))

        numeric_columns: set[int] = set()
        for col_idx, header in enumerate(headers):
            h = header.strip().lower()
            if any(hint in h for hint in numeric_header_hints):
                numeric_columns.add(col_idx)
                continue

            numeric_hits = 0
            text_hits = 0
            for row in table_data[1:]:
                if not isinstance(row, list) or col_idx >= len(row):
                    continue
                cell = str(row[col_idx]).strip()
                if (
                    not cell
                    or cell.startswith("Fecha:")
                    or cell.startswith("Total registros")
                ):
                    continue
                if _is_numeric_value(cell):
                    numeric_hits += 1
                else:
                    text_hits += 1
            if numeric_hits > 0 and numeric_hits >= text_hits:
                numeric_columns.add(col_idx)

        # Width strategy: numeric columns stay compact; text columns use remaining width.
        def _estimate_col_chars(col_idx: int, sample_limit: int = 120) -> int:
            max_len = len(str(headers[col_idx]))
            for row in table_data[1 : min(len(table_data), sample_limit + 1)]:
                if col_idx >= len(row):
                    continue
                cell = str(row[col_idx])
                if cell.startswith("Fecha:") or cell.startswith("Total registros"):
                    continue
                max_len = max(max_len, len(cell))
            return max_len

        col_widths: list[float] = [0.0] * len(headers)
        numeric_total = 0.0
        for col_idx in range(len(headers)):
            if col_idx in numeric_columns:
                chars = _estimate_col_chars(col_idx)
                # Narrow but readable numeric columns.
                width = min(max(18 * mm, chars * 1.8 * mm), 34 * mm)
                col_widths[col_idx] = width
                numeric_total += width

        text_indices = [
            idx for idx in range(len(headers)) if idx not in numeric_columns
        ]
        remaining_width = max(
            doc.width - numeric_total, 24 * mm * max(len(text_indices), 1)
        )
        if text_indices:
            weights = [max(8, _estimate_col_chars(idx)) for idx in text_indices]
            total_weight = sum(weights)
            for idx, weight in zip(text_indices, weights):
                col_widths[idx] = remaining_width * (weight / total_weight)

        # Fallback safety for any uninitialized width.
        for idx in range(len(col_widths)):
            if col_widths[idx] <= 0:
                col_widths[idx] = doc.width / max(len(headers), 1)

        table = Table(table_data, repeatRows=1, hAlign="LEFT", colWidths=col_widths)
        styles_list = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [colors.white, colors.HexColor("#F9FAFB")],
            ),
            ("LINEABOVE", (0, 0), (-1, 0), 1, colors.HexColor("#6B7280")),
            ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#6B7280")),
            ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]

        for row_idx in section_rows:
            styles_list.extend(
                [
                    ("SPAN", (0, row_idx), (-1, row_idx)),
                    ("FONTNAME", (0, row_idx), (-1, row_idx), "Helvetica-Bold"),
                    (
                        "BACKGROUND",
                        (0, row_idx),
                        (-1, row_idx),
                        colors.HexColor("#EEF2F7"),
                    ),
                    (
                        "LINEABOVE",
                        (0, row_idx),
                        (-1, row_idx),
                        0.6,
                        colors.HexColor("#9CA3AF"),
                    ),
                    (
                        "LINEBELOW",
                        (0, row_idx),
                        (-1, row_idx),
                        0.6,
                        colors.HexColor("#9CA3AF"),
                    ),
                ]
            )

        for row_idx in subtotal_rows:
            styles_list.extend(
                [
                    ("SPAN", (0, row_idx), (-1, row_idx)),
                    ("FONTNAME", (0, row_idx), (-1, row_idx), "Helvetica-Bold"),
                    (
                        "TEXTCOLOR",
                        (0, row_idx),
                        (-1, row_idx),
                        colors.HexColor("#111827"),
                    ),
                    (
                        "BACKGROUND",
                        (0, row_idx),
                        (-1, row_idx),
                        colors.HexColor("#F3F4F6"),
                    ),
                    ("ALIGN", (0, row_idx), (-1, row_idx), "RIGHT"),
                ]
            )

        for col_idx in sorted(numeric_columns):
            styles_list.append(("ALIGN", (col_idx, 1), (col_idx, -1), "RIGHT"))

        table.setStyle(TableStyle(styles_list))
        story.append(table)
        doc.build(story, onFirstPage=_draw_page_footer, onLaterPages=_draw_page_footer)
        return buffer.getvalue()

    @staticmethod
    def to_excel(
        headers: list[str],
        rows: list[list[Any]],
        title: str,
        sheet_name: str = "Reporte",
        company_header: dict[str, Any] | None = None,
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        header_fill = PatternFill(
            start_color="2C3E50", end_color="2C3E50", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        row_offset = 1

        # Corporate header rows
        if company_header:
            razon = company_header.get("razon_social", "")
            nombre = company_header.get("nombre_comercial")
            ruc = company_header.get("ruc", "")
            generated_at = company_header.get("generated_at", "")

            ws.merge_cells(
                start_row=row_offset,
                start_column=1,
                end_row=row_offset,
                end_column=len(headers),
            )
            c = ws.cell(row=row_offset, column=1, value=razon)
            c.font = Font(bold=True, size=13)
            c.alignment = Alignment(horizontal="center")
            row_offset += 1

            if nombre:
                ws.merge_cells(
                    start_row=row_offset,
                    start_column=1,
                    end_row=row_offset,
                    end_column=len(headers),
                )
                ws.cell(row=row_offset, column=1, value=nombre).alignment = Alignment(
                    horizontal="center"
                )
                row_offset += 1

            ws.merge_cells(
                start_row=row_offset,
                start_column=1,
                end_row=row_offset,
                end_column=len(headers),
            )
            ws.cell(row=row_offset, column=1, value=f"RUC: {ruc}").alignment = (
                Alignment(horizontal="center")
            )
            row_offset += 1

            ws.merge_cells(
                start_row=row_offset,
                start_column=1,
                end_row=row_offset,
                end_column=len(headers),
            )
            ws.cell(
                row=row_offset, column=1, value=f"Generado: {generated_at}"
            ).alignment = Alignment(horizontal="center")
            row_offset += 1

        # Title row
        ws.merge_cells(
            start_row=row_offset,
            start_column=1,
            end_row=row_offset,
            end_column=len(headers),
        )
        title_cell = ws.cell(row=row_offset, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row_offset].height = 25
        row_offset += 1

        # Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_offset, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws.row_dimensions[row_offset].height = 20
        row_offset += 1

        # Data rows
        for row_idx, row in enumerate(rows, row_offset):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col_idx in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(
                    row_offset - 1, min(len(rows) + row_offset, row_offset + 100)
                )
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_len + 2, 40
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
